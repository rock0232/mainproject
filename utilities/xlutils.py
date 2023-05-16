import openpyxl
import gspread
from pathlib import Path
# import pandas as pd
from oauth2client.service_account import ServiceAccountCredentials


def getRowCount(file, sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return sheet.max_row


def getColumnCount(file, sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return sheet.max_column


def readData(file, sheetName, rownum, columnno):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return sheet.cell(row=rownum, column=columnno).value


def writeData(file, sheetName, rownum, columnno, data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    sheet.cell(row=rownum, column=columnno).value = data
    workbook.save(file)


def get_project_root() -> Path:
    return Path(__file__).parent.parent


# Get data from Google spreadsheet
# Define the scope and credentials
scope = ['https://spreadsheets.google.com/feeds', 'https://www.googleapis.com/auth/drive']
filepath = get_project_root()
credentials = ServiceAccountCredentials.from_json_keyfile_name(f'{filepath}/utilities/credential.json', scope)

# Authenticate and create the client
client = gspread.authorize(credentials)

# Open the Google Sheet
sheet = client.open('B2B Website List').sheet1


# Read data from the Google Sheet and return as a list
def read_sheet():
    data = sheet.get_all_records()
    urllist = []
    for value in data:
        url = (value["Website URL"])
        urllist.append(url)
    return urllist


# read_sheet("Fail")


# Write data to the Google Sheet from a DataFrame
def write_sheet(websitename, value):
    cell = sheet.find(websitename)
    rownum = cell.row
    # sheet.clear()  # Optional: Clear the existing data in the sheet
    # sheet.update(row, col, value)
    sheet.update_cell(rownum, 3, value)

# write_sheet("https://aura24.bet", "PASSE")
# Example usage
# data_to_write = pd.DataFrame({
#     'Name': ['John', 'Jane', 'Alice'],
#     'Age': [25, 30, 35]
# })

# write_sheet(data_to_write)
# data_read = read_sheet()
# print(type(data_read))
